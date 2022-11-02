"""
Utilities.
"""

import os
import io
import re
import sys
import math
import yaml
import inspect
import numbers
import argparse
import numpy as np
import pandas as pd
import urllib.request
from zipfile import ZipFile
from string import Template
import matplotlib.pyplot as plt
from bs4 import BeautifulSoup
from openpyxl import load_workbook

MODULE_DIR = os.path.dirname(os.path.realpath(__file__))
PROJECT_DIR = os.path.abspath(os.path.join(MODULE_DIR, '../..'))
CONFIG_FILES = [
  os.path.join(PROJECT_DIR, 'config.yml'),
  os.path.join(PROJECT_DIR, 'local_config.yml'),
]

# Excel removes some characters from the sheet names and limits their length
MOST_IMPORTANT_PARAMETERS_SHEET = re.sub(r'[\[\]]', '', '[tom] Most important parameters and metrics')[:31]
AJEYA_SHEET = 'Ajeya distribution of automation FLOP'[:31]

#--------------------------------------------------------------------------
# Poor person's mini configuration system
#--------------------------------------------------------------------------

config = None
config_loaded = False

def get_options():
  global config_loaded
  if not config_loaded:
    load_config_files()

  global config
  return config

def get_option(name, default = None):
  return get_options().get(name, default)

def set_option(name, value):
  global config_loaded
  if not config_loaded:
    load_config_files()

  global config
  config[name] = value

def load_config_files():
  configs = []
  for file in CONFIG_FILES:
    if os.path.exists(file):
      with open(file, "r") as f:
        result = yaml.load(f, Loader=yaml.FullLoader)
        if result:
          configs.append(result)

  global config
  config = merge_dicts(configs)
  expand_config_object(config)

  global config_loaded
  config_loaded = True

def merge_dicts(dicts):
  keys = []
  for d in dicts:
    for k in d.keys():
      if k not in keys:
        keys.append(k)

  merged = {}
  for k in keys:
    dicts_with_key = [d for d in dicts if k in d]

    for d in dicts_with_key:
      if type(d[k]) != type(dicts_with_key[0][k]) \
          and not (isinstance(d[k], numbers.Number) and isinstance(dicts_with_key[0][k], numbers.Number)):
        raise Exception(f'Conflict merging dictionaries (key {k})')

    if isinstance(dicts_with_key[0][k], dict):
      merged[k] = merge_dicts([d[k] for d in dicts_with_key])
    else:
      merged[k] = dicts_with_key[-1][k]

  return merged

def expand_config_object(config):
  context = {
    'project_dir': PROJECT_DIR,
  }
  _expand_config_object(config, context, '')

def _expand_config_object(obj, context, path):
  if isinstance(obj, (dict, list)):
    items = obj.items() if isinstance(obj, dict) else enumerate(obj)
    for key, value in items:
      if isinstance(value, (dict, list)):
        _expand_config_object(value, context, f'{path}{key}.')
        continue

      if isinstance(value, str):
        obj[key] = expand_string(value, context)
      context[path + key] = obj[key]

def expand_string(string, context):
  # Awfully hacky
  for key, value in context.items():
    string = string.replace(f'${{{key}}}', str(value))
  return string

#--------------------------------------------------------------------------
# Cached parameter retrieval
#--------------------------------------------------------------------------

cached_input_workbooks = {}
cached_ajeya_dist = None
cached_param_table = None
cached_metrics_table = None
cached_rank_correlations = None
cached_timelines_parameters = None

def set_input_workbook(url):
  global input_workbook
  global cached_input_workbooks
  global cached_param_table
  global cached_metrics_table
  global cached_rank_correlations
  global cached_timelines_parameters

  # If it's a Google sheet url, convert it into an excel export
  export_url = get_export_url_from_gsheet(url)
  if export_url:
    url = export_url

  set_option('input_workbook', url)
  cached_input_workbooks = {}
  cached_param_table = None
  cached_metrics_table = None
  cached_rank_correlations = None
  cached_timelines_parameters = None

def get_export_url_from_gsheet(gsheet_url, format = 'xlsx'):
  """Returns and "export" URL from a Google Sheets url"""
  if re.match(r'^(http|https|file)://', gsheet_url):
    gsheets_pattern = r'https://docs.google.com/spreadsheets/d/([a-zA-Z0-9-_]*)/?'
    m = re.match(gsheets_pattern, gsheet_url)
    if m:
      workbook_id = m.group(1)
      gsheet_url = f'https://docs.google.com/spreadsheets/d/{workbook_id}/export?format={format}'
      return gsheet_url
  return None

def get_workbook(path, format = 'xlsx', data_only = False):
  if re.match(r'^(http|https|file)://', path):
    response = urllib.request.urlopen(get_export_url_from_gsheet(path, format = format))
    workbook = response.read()
  else:
    with open(path, 'rb') as f:
      workbook = f.read()

  if format == 'xlsx':
    # Check it's a valid Excel file
    try:
      load_workbook(io.BytesIO(workbook))
    except Exception:
      raise InvalidExcelError("Error reading the Excel file (you might want to check it's publicly accessible)")

  if format == 'zip':
    # Check it's a valid ZIP file
    try:
      ZipFile(io.BytesIO(workbook))
    except Exception:
      raise InvalidZipError("Error reading the zipped HTML files (you might want to check the workbook is publicly accessible)")

  return workbook

def get_input_workbook(format = 'xlsx'):
  global cached_input_workbooks
  if format not in cached_input_workbooks:
    cached_input_workbooks[format] = get_workbook(get_option('input_workbook'), format)
  return cached_input_workbooks[format]

def get_parameter_table():
  global cached_param_table
  if cached_param_table is None:
    url = get_option('param_table_url')
    if url:
      cached_param_table = get_csv_from_sheet_url(url)
    else:
      # By default we read the table from the omni workbook
      cached_param_table = pd.read_excel(get_input_workbook(), sheet_name = 'Parameters')

    def floatify_series(series):
      def floatify(x):
        if isinstance(x, str):
          if x.endswith('%'):
            x = float(x[:-1])/100
          else:
            x = float(x)
        return x
      return series.apply(floatify)

    cached_param_table['Conservative'] = floatify_series(cached_param_table['Conservative'])
    cached_param_table['Best guess']   = floatify_series(cached_param_table['Best guess'])
    cached_param_table['Aggressive']   = floatify_series(cached_param_table['Aggressive'])

    cached_param_table = cached_param_table.set_index("Parameter id")
    cached_param_table.fillna(np.nan, inplace = True)
  return cached_param_table.copy()

def get_sheet_df_as_rich_text(sheet_name, workbook = None):
  # Unfortunately, we can't just read the rich text of the cells from
  # the standard excel file. We'll have to download all the sheets as
  # HTML pages (zipped) and extract the rich text from the <td> elements.

  if workbook is None:
    workbook = get_input_workbook(format = 'zip')

  zip_file = ZipFile(io.BytesIO(workbook))
  sheet_html = zip_file.read(f'{sheet_name}.html')
  zip_file.close()

  page = BeautifulSoup(sheet_html, 'html.parser')
  table = page.find('table')
  tbody = table.find('tbody')

  df_rows = []

  for row in tbody.find_all('tr'):
    df_row = []
    for cell in row.find_all('td'):
      cell_contents = cell.decode_contents()
      df_row.append(cell_contents)
    df_rows.append(df_row)

  df = pd.DataFrame(df_rows[1:], columns = df_rows[0], index = [r[0] for r in df_rows[1:]])

  return df

def get_sheet_table(sheet_name):
  table = pd.read_excel(get_input_workbook(), sheet_name = sheet_name)
  table = table.set_index(table.columns[0])
  return table

def get_metrics_table():
  global cached_metrics_table
  if cached_metrics_table is None:
    cached_metrics_table = get_sheet_table('Output metrics')
  return cached_metrics_table

def set_parameter_table_url(url):
  global cached_param_table
  set_option('param_table_url', url)
  cached_param_table = None

def get_ajeya_dist():
  global cached_ajeya_dist
  if cached_ajeya_dist is None:
    url = get_option('ajeya_dist_url')
    cols = [0, 1]
    if url:
      cached_ajeya_dist = get_csv_from_sheet_url(url, usecols = cols)
    else:
      # By default we read the distibution from the omni workbook
      cached_ajeya_dist = pd.read_excel(get_input_workbook(), sheet_name = AJEYA_SHEET, usecols = cols)

  return cached_ajeya_dist.copy()

def get_clipped_ajeya_dist(lower_bound):
  ajeya_dist = get_ajeya_dist()

  if lower_bound:
    lower_bound = np.log10(lower_bound)
    i = np.argmax(ajeya_dist.iloc[:, 0] >= lower_bound)
    clip_p = ajeya_dist.iloc[i, 1]

    # Clip...
    ajeya_dist.iloc[i:, 1] -= clip_p
    ajeya_dist.iloc[:i, 1] = 0

    # ... and renormalize
    ajeya_dist.iloc[i:, 1] /= ajeya_dist.iloc[-1, 1]

  return ajeya_dist

def set_ajeya_dist_url(url):
  global cached_ajeya_dist
  set_option('ajeya_dist_url', url)
  cached_ajeya_dist = None

def get_rank_correlations():
  global cached_rank_correlations
  if cached_rank_correlations is None:
    cached_rank_correlations = pd.read_excel(get_input_workbook(), sheet_name = 'Rank correlations', skiprows = 2)
    cached_rank_correlations = cached_rank_correlations.set_index(cached_rank_correlations.columns[0])
  return cached_rank_correlations.copy()

def get_timelines_parameters():
  global cached_timelines_parameters
  if cached_timelines_parameters is None:
    cached_timelines_parameters = pd.read_excel(get_input_workbook(), sheet_name = 'Guess FLOP gap & timelines')
    cached_timelines_parameters = cached_timelines_parameters.set_index(cached_timelines_parameters.columns[0])
  return cached_timelines_parameters.copy()

def get_parameters_colors():
  colors = {}

  workbook = load_workbook(io.BytesIO(get_input_workbook()))
  sheet = workbook['Parameters']
  for row in sheet.iter_rows(min_row = 2, max_col = 1):
    cell = row[0]
    param = cell.value
    if not param: break
    colors[param] = f'#{cell.fill.bgColor.rgb[2:]}' # get rid of alpha

  return colors

def get_parameters_meanings():
  return get_parameter_table()['Meaning'].to_dict()

def get_metrics_meanings():
  return get_metrics_table()['Meaning'].to_dict()

def snake_case_to_human(s):
  words = s.split('_')
  words[0] = words[0].capitalize()
  return " ".join(words)

def get_param_names():
  table = get_parameter_table()['Parameter'].to_dict()
  for k, v in table.items():
    if pd.isnull(v): table[k] = snake_case_to_human(k)
  return table

def get_metric_names():
  table = get_metrics_table()['Metric'].to_dict()
  for k, v in table.items():
    if pd.isnull(v): table[k] = snake_case_to_human(k)
  return table

def get_variable_names():
  table = get_sheet_table('Variables')['Variable'].to_dict()
  for k, v in table.items():
    if pd.isnull(v): table[k] = snake_case_to_human(k)
  return table

def get_parameter_justifications():
  df = get_sheet_df_as_rich_text('Parameters')
  return df['Justification for assumption'].to_dict()

def get_most_important_metrics():
  table = pd.read_excel(get_input_workbook(), sheet_name = MOST_IMPORTANT_PARAMETERS_SHEET)
  metrics = table['Metric id'].dropna().tolist()
  return metrics

def get_most_important_parameters():
  table = pd.read_excel(get_input_workbook(), sheet_name = MOST_IMPORTANT_PARAMETERS_SHEET)
  params = table['Parameter id'].dropna().tolist()
  return params

#--------------------------------------------------------------------------
# CLI parameters
#--------------------------------------------------------------------------

class ArgumentParserWrapper(argparse.ArgumentParser):
  def __init__(self):
    self.abbrev_to_full_name = {}
    super().__init__()

  def add_argument(self, *args, **kwargs):
    if len(args) >= 2:
      abbr = args[0][1:] # get rid of '-'
      full_name = args[1][2:] # get rid of '--'
      self.abbrev_to_full_name[abbr] = full_name

    super().add_argument(*args, **kwargs)

# Some constants
def init_cli_arguments():
  parser = ArgumentParserWrapper()

  parser.add_argument(
    "-o",
    "--output-file",
    default=None,
    help="Path of the output report (absolute or relative to the report directory)"
  )

  parser.add_argument(
    "-d",
    "--output-dir",
    default=None,
    help="Path of the output directory (will be create if it doesn't exist)"
  )

  parser.add_argument(
    "-w",
    "--workbook-url",
    default=None,
    help="Url of the omni workbook"
  )

  parser.add_argument(
    "--t-start",
    type=int,
    default=None,
  )

  parser.add_argument(
    "--t-end",
    type=int,
    default=None,
  )

  parser.add_argument(
    "--t-step",
    type=float,
    default=None,
  )

  parser.add_argument(
    "-p",
    "--param-table-url",
    default=None,
  )

  parser.add_argument(
    "--ajeya-dist-url",
    default=None,
  )

  parser.add_argument(
    "--human-names",
    action='store_true',
    default=None,
    help="Show human names for the parameters and metrics",
  )

  return parser

def handle_cli_arguments(parser):
  # Load arguments from config files (if any)
  caller = inspect.stack()[1]
  module_name = inspect.getmodulename(caller.filename)
  options = get_option(module_name)
  if options:
    clean_options = {}
    for k, v in options.items():
      if k in parser.abbrev_to_full_name:
        k = parser.abbrev_to_full_name[k]
      k = k.replace('-', '_')
      clean_options[k] = v
    parser.set_defaults(**clean_options)

  args = parser.parse_args()

  if args.workbook_url is not None:
    set_input_workbook(args.workbook_url)

  if args.param_table_url is not None:
    set_parameter_table_url(args.param_table_url)

  if args.ajeya_dist_url is not None:
    set_ajeya_dist_url(args.ajeya_dist_url)

  if args.human_names is not None: set_option('human_names', args.human_names)

  if args.t_start is not None: set_option('t_start', args.t_start)
  if args.t_end is not None: set_option('t_end', args.t_end)
  if args.t_step is not None: set_option('t_step', args.t_step)

  return args

#--------------------------------------------------------------------------
# Misc
#--------------------------------------------------------------------------

class InvalidExcelError(Exception):
  pass

class InvalidCsvError(Exception):
  pass

class InvalidZipError(Exception):
  pass


def get_csv_from_sheet_url(url, usecols = None):
  pattern = r'https://docs.google.com/spreadsheets/d/([a-zA-Z0-9-_]*)/.*\bgid\b=([0-9]*)?.*'
  m = re.match(pattern, url)
  if m:
    workbook_id = m.group(1)
    sheet_id = m.group(2)
    url = f'https://docs.google.com/spreadsheets/d/{workbook_id}/export?format=csv&gid={sheet_id}'

  try:
    csv = pd.read_csv(url, usecols = usecols)
  except Exception:
    raise InvalidCsvError("Error reading the sheet (you might want to check it's publicly accessible)")

  return csv

def draw_oom_lines():
  low, high = plt.gca().get_ylim()
  for oom in range(math.floor(np.log10(low)), math.ceil(np.log10(high))):
    plt.axhline(10**oom, linestyle='dotted', color='black')

# Import display in non IPython environments
try:
  from IPython.display import display
except ModuleNotFoundError: 
  # As a fallback, just print
  def display(x):
    print(x)

class Log:
  ERROR_LEVEL = 1
  INFO_LEVEL  = 2
  TRACE_LEVEL = 3

  def __init__(self, level=None):
    self.level = level if (level is not None) else INFO_LEVEL
    self.indentation_level = 0

  def indent(self):
    self.indentation_level += 1

  def deindent(self):
    self.indentation_level -= 1
    if self.indentation_level < 0: self.indentation_level = 0

  def trace(self, *args, **kwargs):
    self.print(log.TRACE_LEVEL, *args, **kwargs)

  def info(self, *args, **kwargs):
    self.print(log.INFO_LEVEL, *args, **kwargs)

  def error(self, *args, **kwargs):
    self.print(log.ERROR_LEVEL, *args, **kwargs)

  def print(self, level, *args, **kwargs):
    if self.level < level: return

    string_buffer = io.StringIO()
    print(*args, **kwargs, file = string_buffer)
    string = string_buffer.getvalue()

    ends_in_newline = len(string) >= 1 and string[-1] == '\n'
    lines = string.splitlines()

    out = sys.stderr if (level == Log.ERROR_LEVEL) else sys.stdout
    for i, line in enumerate(lines):
      end = '' if (i == len(lines) - 1 and not ends_in_newline) else '\n'
      print((' ' * self.indentation_level) + line, file = out, end = end)

log = Log(level=Log.INFO_LEVEL)
