from . import utils
from .utils import log
from .opmodel import SimulateTakeOff

import io
import urllib.request
import numpy as np
import pandas as pd
from openpyxl import load_workbook

class ScenarioRunner:
  def __init__(self):
    self.groups = None
    self.metrics = None

  def simulate_all_scenarios(self):
    parameter_table = get_parameter_table()
    timelines_parameters = get_timelines_parameters()

    groups = []

    scenarios = self.simulate_scenario_group(parameter_table)
    groups.append(ScenarioGroup('normal', scenarios, parameter_table))

    for timeline in timelines_parameters:
      parameters = timelines_parameters[timeline]

      parameter_table = parameter_table.copy()

      parameter_table.loc['full_automation_requirements_training', 'Conservative'] = parameters['Full automation requirements']
      parameter_table.loc['full_automation_requirements_training', 'Best guess']   = parameters['Full automation requirements']
      parameter_table.loc['full_automation_requirements_training', 'Aggressive']   = parameters['Full automation requirements']

      parameter_table.loc['flop_gap_training', 'Conservative'] = parameters['Long FLOP gap']
      parameter_table.loc['flop_gap_training', 'Best guess']   = parameters['Med FLOP gap']
      parameter_table.loc['flop_gap_training', 'Aggressive']   = parameters['Short FLOP gap']

      log.indent()
      scenarios = self.simulate_scenario_group(parameter_table)
      groups.append(ScenarioGroup(timeline, scenarios, parameter_table))
      log.deindent()

      log.info('')

    self.groups = groups

    return groups

  def simulate_scenario_group(self, parameter_table, simulation_type = 'compare'):
    # Define parameters
    med_params = {
      parameter : row['Best guess']
      for parameter, row in parameter_table.iterrows()
    }

    if simulation_type == 'compare':

      low_params = {
          parameter : row['Conservative'] \
                      if not np.isnan(row['Conservative']) \
                      and row["Compare?"] == 'Y'
                      else row['Best guess']\
          for parameter, row in parameter_table.iterrows()
      }

      high_params = {
          parameter : row['Aggressive'] \
                      if not np.isnan(row['Aggressive']) \
                      and row["Compare?"] == 'Y'
                      else row['Best guess']\
          for parameter, row in parameter_table.iterrows()
      }

      low_value = 'Conservative'
      med_value = 'Best guess'
      high_value = 'Aggressive'

    else:
      row = parameter_table.loc[exploration_target, :]

      low_params = med_params.copy()
      low_params[exploration_target] = row['Conservative']

      high_params = med_params.copy()
      high_params[exploration_target] = row['Aggressive']

      low_value = row['Conservative']
      med_value = row['Best guess']
      high_value = row['Aggressive']

    # Run simulations
    log.info('Running simulations...')

    log.info('  Conservative simulation')
    low_model = SimulateTakeOff(**low_params)
    low_model.run_simulation()

    log.info('  Best guess simulation')
    med_model = SimulateTakeOff(**med_params)
    med_model.run_simulation()

    log.info('  Aggressive simulation')
    high_model = SimulateTakeOff(**high_params)
    high_model.run_simulation()

    # hacky
    if self.metrics is None:
      self.metrics = [MetricDescription(name, name) for name in med_model.get_takeoff_metrics()]

    scenarios = [
      Scenario('Conservative', low_model,  low_params),
      Scenario('Best guess',   med_model,  med_params),
      Scenario('Aggressive',   high_model, high_params),
    ]

    return scenarios

class ScenarioGroup:
  def __init__(self, name, scenarios, parameter_table):
    self.name = name
    self.scenarios = scenarios
    self.parameter_table = parameter_table

  def __getitem__(self, index):
    return self.scenarios[index]

  def __len__(self):
    return len(self.scenarios)

class Scenario:
  def __init__(self, name, model, params):
    self.name    = name
    self.model   = model
    self.params  = params
    self.metrics = [Metric(name, value[0]) for name, value in model.get_takeoff_metrics().iteritems()]

class MetricDescription:
  def __init__(self, name, meaning):
    self.name = name
    self.meaning = meaning

class Metric:
  def __init__(self, name, value):
    self.name = name
    self.value = value

#--------------------------------------------------------------------------
# Parameters (with a cache)
#--------------------------------------------------------------------------

cached_omni_excel = None

def get_omni_excel():
  global cached_omni_excel
  if cached_omni_excel is None:
    response = urllib.request.urlopen('https://docs.google.com/spreadsheets/d/1r-WxW4JeNoi_gCMc5y2iTlJQnan_LLCF5s_V4ZDDMkI/export?format=xlsx')
    cached_omni_excel = response.read()
  return cached_omni_excel

def get_parameter_table():
  parameter_table = pd.read_excel(get_omni_excel(), sheet_name = 'Parameters')
  parameter_table = parameter_table.set_index("Parameter")
  parameter_table.fillna(np.nan, inplace = True)
  return parameter_table

def get_timelines_parameters():
  timelines_parameters = pd.read_excel(get_omni_excel(), sheet_name = 'Guess FLOP gap & timelines')
  timelines_parameters = timelines_parameters.set_index(timelines_parameters.columns[0])
  return timelines_parameters

def get_parameters_meanings():
  param_table = get_parameter_table()
  return param_table['Meaning'].to_dict()

def get_parameters_colors():
  colors = {}

  workbook = load_workbook(io.BytesIO(get_omni_excel()))
  sheet = workbook['Parameters']
  for row in sheet.iter_rows(min_row = 2, max_col = 1):
    cell = row[0]
    param = cell.value
    if not param: break
    colors[param] = f'#{cell.fill.bgColor.rgb[2:]}' # get rid of alpha

  return colors

def get_metrics_meanings():
  metric_meanings = pd.read_excel(get_omni_excel(), sheet_name = 'Output metrics')
  metric_meanings = metric_meanings.set_index(metric_meanings.columns[0])
  return metric_meanings['Meaning'].to_dict()
